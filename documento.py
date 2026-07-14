import subprocess
from subprocess import Popen
import os
from fitz import fitz, Rect
import openpyxl
import threading
from constantes import constantes
from General import General
from PIL import Image, ImageTk
import math
import random
import time
from ventana_sec import ventana_secundaria
from openpyxl.styles import Font,Alignment
from openpyxl_image_loader import SheetImageLoader
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side

#for Operations of Read and Write Over Documents
class documento:
     
     thread_object=None
     result=None
     estatus=0
     queue=[]
     
     #Verify if the Thread is Finished each 1000 ms
     @classmethod
     def check(cls,root,thread):
         root.after(1000,cls.check_ifDone,root,thread)
      
     #Verify if the Thread is Finished and Execute the Required Operations 
     @classmethod
     def check_ifDone(cls,root,thread,comp=None):
     
        if(thread.is_alive()):
               cls.check(root,thread)
        else:
           if(cls.result!=None):
              from tiempo import tiempo
              from conexion_bd import conexion_bd
              from event_manager import Event_manager 
              if(type(cls.result).__name__=="list"):
                 if(cls.result!=[]):
                     if(cls.result[0]=="data formato"):
                          data=[]
                          temp_dat=cls.result
                          for i in range(0,len(temp_dat[2])):
                              data.append(temp_dat[2][i])
                          if(data!=[]):
                            Event_manager.user.recibe_data_process(cls.result[3])
                            Event_manager.set_comp_values(cls.result[1],data)          
                     elif(cls.result[0]=="modificar excel"):
                           Event_manager.finish_update_formato(cls.result[1])               
              elif(cls.result==constantes.REQUEST_RESULT_WRITE_CSV_SUCCESS):
                    time_object=tiempo()
                    Event_manager.user.add_action_historial(["respaldar BD",time_object.get_tiempo()])
                    Event_manager.activar_element("cargando",False,True)
                    conexion_bd.set_tabla(constantes.TABLA_REPORTE)
                    id_hist=conexion_bd.generate_id(True,constantes.CLAVE_REPORTE)
                    data_hist=[ id_hist,Event_manager.user.user,time_object.get_fecha(),time_object.get_tiempo(),"base de datos","respaldo","",time_object.get_fecha()]
                    conexion_bd.add_data(data_hist)
                    General.show_message("respaldo creado exitosamente","Respaldo Generado")
              elif(cls.result==constantes.REQUEST_RESULT_READ_CSV_SUCCESS):
                  time_object=tiempo()
                  Event_manager.user.add_action_historial(["restaurar BD",time_object.get_tiempo()])
                  Event_manager.activar_element("cargando",False,True)
                  conexion_bd.set_tabla(constantes.TABLA_REPORTE)
                  id_hist=conexion_bd.generate_id(True,constantes.CLAVE_REPORTE)
                  data_hist=[ id_hist,Event_manager.user.user,time_object.get_fecha(),time_object.get_tiempo(),"base de datos","restaurar","",time_object.get_fecha()]
                  conexion_bd.add_data(data_hist)
                  General.show_message("restauracion del respaldo realizada exitosamente","Restauracion Exitosa")
              elif(cls.result==constantes.REQUEST_RESULT_GENERATE_DOCUMENT_WITH_MESSAGE):
                  General.show_message("documentos generados","reportes creados")
              elif(cls.result==constantes.REQUEST_RESULT_DOWNLOAD_SUCCESS):
                  General.show_message("documentos descargados exitosamente","Descarga Exitosa")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_WRITE_CSV):
                  General.show_error("error el crear respaldo","Error al Crear Respaldo")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_READ_CSV):
                  Event_manager.activar_element("cargando",False,True)
                  General.show_error("error en la lectura del respaldo","Error al Leer Respaldo")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_NO_REFERENCE_EXCEL):
                  General.show_error("error en no hay contenido de refernecia para escribir en el formato","Referencia Del formato Invalida")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL):
                  General.show_error("error referencia no encontrada","Referencia Del Formato Invalida")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_ACCESS):
                  General.show_error("error de acceso a la data,asegurese que el documento no esta abierto","Acceso Invalido")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_GET_DATA):
                 General.show_error("error obteniendo recurso","error al encontrar recurso")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_FOUND_END_ROW):
                 General.show_error("error imposible encontrar fila final del formato","Formato No Modificable")
              elif(cls.result==constantes.REQUEST_RESULT_ERROR_INVALID_CONTENT):
                 General.show_error("error insuficiente contenido para escribir el documento","Contenido Invalido")
              cls.result=None
              if(cls.queue!=[]):
                  last=len(cls.queue)-1
                  dat_queue=cls.queue[last]
                  temp=[]
                  for i in range(0,last):
                      temp.append(cls.queue[i])
                  cls.queue=[]
                  cls.queue=temp
                  cls.request(dat_queue[0],dat_queue[1],dat_queue[2],dat_queue[3])
      
     #Send a Request for write or read a File
     @classmethod
     def request(cls,root,file_name,request_type,content=[None,None,None],open_file=False):
           
        if(cls.estatus==0):
            if(request_type==constantes.REQUEST_READ_EXCEL):
                cls.thread_object=threading.Thread(target=cls.read_excell,args=(file_name,content))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object)
            elif(request_type==constantes.REQUEST_WRITE_CSV):
                cls.thread_object=threading.Thread(target=cls.write_CSV,args=(None,content))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object)
            elif(request_type==constantes.REQUEST_READ_CSV):
                cls.thread_object=threading.Thread(target=cls.read_CSV,args=(file_name,content))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object)
            elif(request_type==constantes.REQUEST_MODIFIC_EXCEL):
                cls.thread_object=threading.Thread(target=cls.modif_excell,args=(file_name,content))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object)    
            elif(request_type==constantes.REQUEST_WRITE_EXCEL_FROM_EXISTENT_FORMAT):
                cls.thread_object=threading.Thread(target=cls.write_excell_from_formato,args=(file_name,content,open_file))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object)  
            elif(request_type==constantes.REQUEST_DOWNLOAD):
                cls.thread_object=threading.Thread(target=cls.download,args=(file_name,content,open_file))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object)
            elif(request_type==constantes. REQUEST_WRITE_NOTAS_CERTIFICADAS):
                cls.thread_object=threading.Thread(target=cls.write_notas_certific,args=(file_name,content))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object) 
            elif(request_type==constantes.REQUEST_MODIFIC_PDF):
                cls.thread_object=threading.Thread(target=cls.modif_pdf,args=(file_name,content))
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object) 
            elif(request_type==constantes.REQUEST_WRITE_CRONOGRAM):
                cls.thread_object=threading.Thread(target=cls.write_cronograma,args=(file_name,content,open_file))                
                cls.thread_object.start() 
                cls.check_ifDone(root,cls.thread_object) 
            return 0
        else:
            cls.queue.append([root,file_name,request_type,content])  

     #Download a Document
     @classmethod 
     def download(cls,file_name,content,open_file):
        import requests     
        response=requests.get(content[0])
        if(response.status_code>400):
           cls.estatus=0
           cls.result=constantes.REQUEST_RESULT_ERROR_GET_DATA
           return constantes.REQUEST_RESULT_ERROR_GET_DATA
        data=response.content 
        file_name=file_name.replace("ñ","n")
        file_name=file_name.replace("Ñ","N")
        temp_file=open(file_name,"wb")
        temp_file.write(data)
        temp_file.close()
        if(open_file):
            #Open doc in Navigator
            os.startfile(file_name)
            cls.estatus=0
            cls.result=constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
            return constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
        else:
           #Only Download
           cls.estatus=0
           cls.result=constantes.REQUEST_RESULT_DOWNLOAD_SUCCESS
           return constantes.REQUEST_RESULT_DOWNLOAD_SUCCESS
     
     #Modific and Save a Pdf Doc
     @classmethod  
     def modif_pdf(cls,file_name,content):
        cls.estatus=0
        path=file_name[0]
        doc=fitz.open(stream = file_name[1], filetype="pdf")
        num_pages=len(content)
        for temp_pag in range(0,num_pages):
             page=doc[temp_pag]
             replace=content[temp_pag][1]
             num_replaces=content[temp_pag][0]
             if(num_replaces>0):
               for i in range(0,num_replaces):
                 old=replace[i][0]
                 better=replace[i][1]
                 size_l=replace[i][2]
                 bld=replace[i][3]
                 is_bold=False
                 if(bld=="bold"):
                     is_bold=True
                 hits = page.search_for(old)
                 if(hits!=[]):
                    if(old!="FOTO:"):
                      for rect in hits:
                         if(is_bold==False):
                             annot = page.add_redact_annot(rect, text=better,fontsize=size_l)
                         else:
                             annot = page.add_redact_annot(rect, text=better,fontsize=size_l,fontname="courier-bold")
                      page.apply_redactions()
        path=path.replace("ñ","n")
        path=path.replace("Ñ","N")
        doc.save(path, garbage=3, deflate=True)
        os.startfile(ruta)
        cls.estatus=0
        cls.result=constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
        return 0
 
     #Modify Excel File
     @classmethod          
     def modif_excell(cls,file_name,content):
        if(len(content)<4):
             General.show_error("contenido incompleto","datos incompletos")
        values=content[0]
        reference=content[1]
        row_index=str(content[2])
        try:
          from io import BytesIO
          temp_file=BytesIO(file_name[0])         
          cls.estatus=50 
          excl= openpyxl.load_workbook(temp_file,data_only=True)
          pages= excl.active
          found=False
          borde_row=-1
          letter_index=-1
          columns=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
          end_col=len(columns)
          #get the reference data ( row and letter/column Index in document)
          for i in range(0,end_col):
              celda=pages[columns[i]+row_index]
              val=None
              if(celda!=None):
                 val=celda.value
              if(val!=None and val==reference):
                found=True
                letter_index=i+1
                for row in range(int(row_index)+1,100):
                    celda_temp=pages[columns[i]+str(row)]  
                    try:
                        if(borde_row==-1):                  
                           if(celda_temp.border.left.style is None and celda_temp.border.right.style is None ):
                               if(celda_temp.border.bottom.style is None and celda_temp.border.top.style is None):
                                   borde_row=row-1
                    except:
                        borde_row=row-1                                          
              if(found):
                  break
                  
          if(found==False or letter_index==-1):
              cls.estatus=0
              cls.result=constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL
              return -9 
          if(borde_row==-1):
            cls.estatus=0
            cls.result=constantes.REQUEST_RESULT_ERROR_FOUND_END_ROW
            return constantes.REQUEST_RESULT_ERROR_FOUND_END_ROW 
            
          num_campos=len(values)
          count=1
          for j in range(letter_index,end_col):
            if(count<num_campos):
                celda=pages[columns[j]+row_index] 
                val=""
                if(celda!=None):
                   val=celda.value
                if(val==None or val=="" or val==" "):
                    #Modify Border
                    celda.alignment=Alignment(horizontal="center")   
                    borde=Border(left=Side(style='thin'),right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    celda.border=borde
                    if(j>=1):
                        col_prev=pages[columns[j-1]+row_index]
                        if(col_prev!=None):
                            if(col_prev.font!=None):
                               previous_font=pages[columns[j-1]+row_index].font
                               celda.font=Font(name=previous_font.name, size=previous_font.sz,bold=previous_font.b)                      
                            else:
                               celda.font=Font(name="Calibri", size=11,bold=True)                      
                    for k in range(int(row_index)+1,borde_row+1):
                        celda_temp=pages[columns[j]+str(k)] 
                        celda_temp.alignment=Alignment(horizontal="center")   
                        top_border=False
                        if(j>=1):
                            previous_cell=pages[columns[j-1]+str(k)] 
                            if(previous_cell!=None):
                                valid_border=False
                                if(previous_cell.border!=None):
                                    if(previous_cell.border.top!=None):
                                         valid_border=True
                                if(valid_border):
                                    if(previous_cell.border.top.style is not None):
                                       top_border=True
                        if(k<borde_row):
                            if(top_border==False):
                                borde_t=Border(left=Side(style='thin'),right=Side(style='thin'))
                                celda_temp.border=borde_t 
                            else:
                                 borde_t=Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))
                                 celda_temp.border=borde_t 
                        else:
                            borde_t=Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'))
                            celda_temp.border=borde_t 
                   
                #Modify Value of Cell                       
                pages[columns[j]+row_index]=values[count]
                count+=1
            else:
                #Remove border from cells irrelevants
                initial=int(row_index)
                if(int(row_index)>1):
                    initial=initial-1
                for col in range(initial,borde_row+1):
                    if(col>=int(row_index)):
                      pages[columns[j]+str(col)]=""
                      pages[columns[j]+str(col)].border=Border(left=Side(style=None),right=Side(style=None),bottom=Side(style=None),top=Side(style=None))
                    else:        
                      pages[columns[j]+str(col)].border=Border(left=Side(style=None),right=Side(style=None),bottom=Side(style=None),top=Side(style=None))          
          path_temp=constantes.FOLDER_DOCUMENTS+file_name[1]
          path_temp=path_temp.replace("Ñ","N")
          path_temp=path_temp.replace("ñ","n") 
          excl.save(path_temp) 
          excl.close()
          resultado=["modificar excel",path_temp]
          cls.estatus=0
          cls.result=resultado
        except:
            cls.estatus=0
            cls.result=constantes.REQUEST_RESULT_ERROR_ACCESS
            return constantes.REQUEST_RESULT_ERROR_ACCESS
          
     #Replace the The Indicated Notes in the required cells and show the document of 'Certified Notes' 
     @classmethod 
     def write_notas_certific(cls,file_name,content):

        if(len(content)<5):     
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_INVALID_CONTENT
             return  constantes.REQUEST_RESULT_ERROR_INVALID_CONTENT            
        reference_cells=content[2]
        col_count=content[3]
        data_replace=content[4]
        num_replaces=len(content[4])
        message_on_End=content[5]
        if(reference_cells==[] or col_count==-1):
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_NO_REFERENCE_EXCEL
             return constantes.REQUEST_RESULT_ERROR_NO_REFERENCE_EXCEL             
        try:                         
            from io import BytesIO
            temp_file=BytesIO(file_name[1])         
            cls.estatus=50 
            excl_f = openpyxl.load_workbook(temp_file,data_only=True)
            pass_sheet="super password"
            pages_file = excl_f.active   
            pages_file.protection.sheet = True
            pages_file.protection.password =pass_sheet
            pages_file.protection.enable() 
            columns=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
            end_col=len(columns)
         
            #Read a Max of 100 Rows
            found=0
            reference_indexs=[] 

            #Replace cell Values with the Notes of Student            
            for i in range(1,100):
               for j in range(0,end_col):             
                  celda=pages_file[columns[j]+str(i)]
                  val=None
                  if(celda!=None):
                    val=celda.value
                  if(val!=None  and num_replaces>0):                    
                      for k in range(0,num_replaces):            
                        if(val==data_replace[k][0]):
                            data_replace[k][1]=data_replace[k][1].upper()          
                            pages_file[columns[j]+str(i)]=val+" "+data_replace[k][1]
                      if(found<5):
                        for k in range(0,len(reference_cells)):
                          if(val==reference_cells[k]):
                            temp_reference=[]
                            temp_reference.append(str(i+2))
                            temp_reference.append(columns[j])
                            temp_reference.append(columns[j+1])
                            temp_reference.append(val[0])
                            reference_indexs.append(temp_reference)
                            found=found+1      
            if(found<=0): 
               cls.estatus=0
               cls.result=constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL
               return constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL  
            for rf in range(0,len(reference_indexs)):
                refe_actual=reference_indexs[rf]
                row_count=content[0]
                row_init=0
                if(refe_actual!=[]):
                  row_init=int(refe_actual[0])
                  for dat in range(0,len(content[1])):
                     if(content[1][dat][0]==refe_actual[3]):
                        califs=content[1][dat]
                        for cal in range(1,len(califs)):
                           area=califs[cal][0]
                           nota=califs[cal][1]                           
                           celda_index=refe_actual[1]+str(row_init+(cal-1))     
                           pages_file[celda_index]=area.upper()
                           celda_index=refe_actual[2]+str(row_init+(cal-1))     
                           pages_file[celda_index]=nota.upper()    
            from openpyxl.workbook.protection import WorkbookProtection
            protect_pass='super-secret-password'
            excl_f.security=WorkbookProtection(workbookPassword = protect_pass, lockStructure = True)
            file_name[0]=file_name[0].replace("ñ","n")
            file_name[0]=file_name[0].replace("Ñ","N")
            excl_f.save(file_name[0])
            excl_f.close()      
            os.startfile(file_name[0])              
            cls.estatus=0
            cls.result=constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
        except:
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_ACCESS
             return constantes.REQUEST_RESULT_ERROR_ACCESS  
      
     #Write a Excel Document from a Existent Format of Excel 
     @classmethod 
     def write_excell_from_formato(cls,file_name,content,open_file=False):
        
        if(len(content)<5):          
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_INVALID_CONTENT
             return constantes.REQUEST_RESULT_ERROR_INVALID_CONTENT           
        reference_cell=content[2]
        col_count=content[3]
        data_replace=content[4]
        num_replaces=len(content[4])
        message_on_End=content[5]
        if(reference_cell=="" or col_count==-1):
            cls.estatus=0
            cls.result=constantes.REQUEST_RESULT_ERROR_NO_REFERENCE_EXCEL
            return constantes.REQUEST_RESULT_ERROR_NO_REFERENCE_EXCEL
        try:     
            from io import BytesIO
            import requests
            pass_sheet="super password"
            temp_file=BytesIO(file_name[1])         
            cls.estatus=50 
            excl_f = openpyxl.load_workbook(temp_file,data_only=True)
            pages_file = excl_f.active   
            pages_file.protection.sheet = True
            pages_file.protection.password =pass_sheet
            pages_file.protection.enable()
            columns=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH']
            end_col=len(columns)
            found=False
            reference_indexs=[]
             
            #Read a Max of 100 Rows
            for i in range(1,100):
               for j in range(0,end_col):
                  celda=pages_file[columns[j]+str(i)]
                  val=None
                  if(celda!=None):
                    val=celda.value
                  if(val!=None and num_replaces>0 ):
                      for k in range(0,num_replaces):                    
                         if(val==data_replace[k][0]):
                            cls.replace_cell_of_format(val,data_replace[k][1],columns[j-1],columns[j],i,pages_file) 
                      if(val==reference_cell and found==False):
                          reference_indexs.append(str(i+1))
                          for k in range(0,col_count):
                            reference_indexs.append(columns[j+(1+k)])
                          found=True
            if(found==False):
               cls.estatus=0
               cls.result=constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL
               return  constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL  
            
            row_count=content[0]
            row_init=0
            if(reference_indexs!=[]):
                row_init=int(reference_indexs[0])
            for row in range(0,row_count):
                for col in range(0,col_count):
                   celda_index=reference_indexs[col+1]+str(row_init+row)
                   pages_file[celda_index]=content[1][row][col].upper()      
          

            file_name[0]=file_name[0].replace("ñ","n")
            file_name[0]=file_name[0].replace("Ñ","N")          
            from openpyxl.workbook.protection import WorkbookProtection
            protect_pass='super-secret-password'
            excl_f.security=WorkbookProtection(workbookPassword = protect_pass, lockStructure = True)  
            if(os.path.exists(file_name[0])):
               os.remove(file_name[0])
            excl_f.save(file_name[0])
            excl_f.close()           
            if(open_file): 
               os.startfile(file_name[0])              
            cls.estatus=0
            
            if(message_on_End):
               cls.result=constantes.REQUEST_RESULT_GENERATE_DOCUMENT_WITH_MESSAGE
               return constantes.REQUEST_RESULT_GENERATE_DOCUMENT_WITH_MESSAGE
            else:
               cls.result=constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
               return constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
        except:            
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_ACCESS
             return constantes.REQUEST_RESULT_ERROR_ACCESS 
     
     #replace value of a cell in a format
     @classmethod
     def replace_cell_of_format(cls,value_cell,value_replace,preview_column,column,index_row,page):
        from io import BytesIO
        import requests
        if(value_cell=="FOTO:"):
            page[column+str(index_row)]=""
            if(value_replace!="not found"):                                        
                try:
                    response=requests.get(value_replace)
                    if(response.status_code>=200 and response.status_code<400):
                        data_img=response.content
                        img = openpyxl.drawing.image.Image(BytesIO(data_img))
                        img.width=75
                        img.height=90
                        img.anchor = column+str(index_row-1)
                        page.add_image(img)
                    else:
                        page[column+str(index_row+1)]="FOTO"
                except:
                    page[column+str(index_row+1)]="FOTO"
            else:
                page[column+str(index_row+1)]="FOTO"
        else:
            valor=value_cell
            left_cell=False
            if(valor=="CEDULA_R:"):
                valor="CEDULA:"
            elif(valor=="NN"):
                valor=""
            elif(valor=="DIRECTOR"):
                 valor=""
            elif(valor=="fecha_e"):
                valor=""
            elif(valor=="fecha_v"):
                valor=""
            elif(valor=="NOMBRES_C"):
                valor=""
            elif(valor=="CARGO_C"):
                valor=""
            elif(valor=="CEDULA_C"):
                valor=""
            elif(valor=="CEDULA_DIRE"):
                valor=""
            elif(valor.startswith("Area N")==True):
                valor=""
            elif(valor=="Directivo"):
                valor=""
            elif(valor=="CI dire"):
                valor="cedula: "
            elif(valor=="code"):
                valor=""
            elif(valor=="en el:"):
                valor="en el "
            elif(valor=="/"):
                valor=""
            elif(valor=="copia de Cedula de Identidad"):
                left_cell=True
            elif(valor=="Copia de Partida de Nacimiento Original"):
                left_cell=True
            elif(valor=="partida de nacimiento Original"):
                left_cell=True
            elif(valor=="Boleta del periodos escolar anterior de ser neceario"):
                left_cell=True
            elif(valor=="Notas Cerificadas"):
                left_cell=True
            elif(valor=="Carta de Residencia"):
                left_cell=True
            elif(valor=="Copia de Cedula"):
                left_cell=True
            elif(valor=="2 Fotos"):
                left_cell=True
            elif(valor=="2 fotos del estudiante"):
                left_cell=True
            if(valor!="Aguas Calientes," and valor!="a los" and valor!="en el " and valor!="Correo del Representante:"):    
                value_replace=value_replace.upper()
            if(left_cell==True):
                page[preview_column+str(index_row)]=value_replace
            else:
                page[column+str(index_row)]=valor+" "+value_replace
      
      #write the data of Cronogram in  a Excel File      
     @classmethod 
     def write_cronograma(cls,file_name,content,open_file=False):
        if(len(content)<5):
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_INVALID_CONTENT
             return constantes.REQUEST_RESULT_ERROR_INVALID_CONTENT   
        reference_cell=content[2]
        col_count=content[3]
        data_replace=content[4]
        message_on_End=content[5]
        if(reference_cell=="" and col_count==-1):
            cls.estatus=0
            cls.result=constantes.REQUEST_RESULT_ERROR_NO_REFERENCE_EXCEL
            return constantes.REQUEST_RESULT_ERROR_NO_REFERENCE_EXCEL 
        try:
            from io import BytesIO
            import requests
            temp_file=BytesIO(file_name[1])         
            cls.estatus=50 
            excl_f = openpyxl.load_workbook(temp_file,data_only=True) 
            for index_page in range(0,3):
                 pages_file = excl_f.worksheets[index_page] 
                 pass_sheet="super password"  
                 pages_file.protection.sheet = True
                 pages_file.protection.password =pass_sheet
                 pages_file.protection.enable()                 
                 num_replaces=len(data_replace[index_page])
                 columns=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH']
                 end_col=len(columns)
                 found=False
                 reference_indexs=[]
                 #Read a Max of 100 Rows
                 for i in range(1,100):
                   for j in range(0,end_col):
                      celda=pages_file[columns[j]+str(i)]
                      val=None
                      if(celda!=None):
                        val=celda.value
                      if(val!=None and num_replaces>0 ):
                          
                        for k in range(0,num_replaces):
                            if(val==data_replace[index_page][k][0]):
                                valor=val
                                if(valor=="xperiodox"):
                                    valor=""
                                elif(valor=="NN"):
                                    valor=""
                                elif(valor=="año_e"):
                                     valor=""
                                elif(valor=="cierre_e"):
                                    valor=""                                    
                                data_replace[index_page][k][1]=data_replace[index_page][k][1].upper()
                                pages_file[columns[j]+str(i)]=valor+" "+data_replace[index_page][k][1]
                        if(val==reference_cell and found==False and content[1][index_page]!=[]):                             
                            reference_indexs.append(str(i+1))
                            for k in range(0,col_count):
                                reference_indexs.append(columns[j+(1+k)])
                            found=True
                            row_count=content[0][index_page]
                            row_init=0
                            if(reference_indexs!=[]):                                
                                row_init=int(reference_indexs[0])
                                for row in range(0,row_count):
                                    for col in range(0,col_count):
                                        celda_index=reference_indexs[col+1]+str(row_init+row)           
                                        pages_file[celda_index]=content[1][index_page][row][col].upper()            
            from openpyxl.workbook.protection import WorkbookProtection
            protect_pass='super-secret-password'
            excl_f.security=WorkbookProtection(workbookPassword = protect_pass, lockStructure = True)
            excl_f.active=excl_f.worksheets[0]      
            file_name[0]=file_name[0].replace("ñ","n")
            file_name[0]=file_name[0].replace("Ñ","N")          
            excl_f.save(file_name[0])
            excl_f.close()                    
            
            if(open_file): 
               os.startfile(file_name[0])              
            cls.estatus=0
            if(message_on_End):
               cls.result=constantes.REQUEST_RESULT_GENERATE_DOCUMENT_WITH_MESSAGE
               return constantes.REQUEST_RESULT_GENERATE_DOCUMENT_WITH_MESSAGE
            else:
               cls.result=constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
               return constantes.REQUEST_RESULT_OPERATION_SUCCEESS_NO_MESSAGE
        except:            
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_ACCESS
             return constantes.REQUEST_RESULT_ERROR_ACCESS  
          
         
     #Read a Excel File         
     @classmethod     
     def read_excell(cls,file_name,content):        
       try:
          from io import BytesIO
          temp_file=BytesIO(file_name)         
          cls.estatus=50 
          excl = openpyxl.load_workbook(temp_file,data_only=True)
          hojas = excl.active                          
          cols=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
          end_col=len(cols)
          row_end=100          
          destino=content[1]
          reference=content[2]          
          values=["data formato",destino]
          found=False
          index=[-1,-1]
          for i in range(1,row_end):
             for j in range(0,end_col):              
                celda=hojas[cols[j]+str(i)]
                if(celda!=None):
                  val=celda.value
                  if(val!=None ):
                    if(val==reference and found==False):
                      found=True
                      index=[j,i]                    
             if(found):
                break                
          if(found==False):
            cls.estatus=0
            cls.result=constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL
            return constantes.REQUEST_RESULT_ERROR_REFERENCE_EXCEL  
          values_cells=[]
          initial_index=""
          for row in range(index[0],end_col):
              cell_index=cols[row]+str(index[1])
              celda=hojas[cell_index]
              if(initial_index==""):
                 initial_index=index[1]
              if(celda!=None):
                val=celda.value
                if(val!=None ):
                   values_cells.append(str(val))
          values.append(values_cells) 
          values.append([reference,initial_index])            
          cls.estatus=0
          cls.result=values
       except:
           cls.estatus=0
           cls.result=constantes.REQUEST_RESULT_ERROR_ACCESS
           return constantes.REQUEST_RESULT_ERROR_ACCESS 
     
     #Write a CSV File with the Tables data from Data Base  
     @classmethod 
     def write_CSV(cls,file_name,content):
         num_tablas=content[0]
         porcent=0       
         from event_manager import Event_manager 
         from conexion_bd import conexion_bd 
         for i in range(0,num_tablas):
           if(i>0):
              porcent=(int((float(i/num_tablas))*100))
              Event_manager.set_comp_values("cargando",f"Escribiendo Datos: {str(porcent)} %")
           else:
              Event_manager.set_comp_values("cargando","Escribiendo Datos: 0 %")
           tabla_name=content[1][i]
           filename=content[2]+tabla_name+".csv"
           conexion_bd.set_tabla(tabla_name)
           data_tabla=conexion_bd.get_allData(None,None)
           line=""
           file_tabla=None
           try:
             file_tabla=open(filename, "w")
             line_header=""
             for j in range(0,len(content[3][i])):
                 if(j!=0):
                    line_header+=';'
                 line_header+=content[3][i][j]
             file_tabla.write(line_header+"\n")            
             for j in range(0,len(data_tabla)):
                line=""
                for k in range(0,len(data_tabla[j])):
                   if(k!=0):
                       line+=';'
                   line+=data_tabla[j][k]               
                file_tabla.write(line+"\n")
             file_tabla.close() 
           except:
             if(file_tabla!=None):
                 file_tabla.close()
             cls.estatus=0
             cls.result=constantes.REQUEST_RESULT_ERROR_WRITE_CSV
             return constantes.REQUEST_RESULT_ERROR_WRITE_CSV  
         cls.estatus=0
         cls.result=constantes.REQUEST_RESULT_WRITE_CSV_SUCCESS
         return constantes.REQUEST_RESULT_WRITE_CSV_SUCCESS
       
     #Read the data of Csv Files to Restore Data Base      
     @classmethod       
     def read_CSV(cls,file_name,content):
         num_tablas=content[0]
         porcent=0
         from event_manager import Event_manager 
         from conexion_bd import conexion_bd
         conexion_bd.set_foregein_check(False)
         for i in range(0,num_tablas):
            if(i>0):
              porcent=(int((float(i/num_tablas))*100))
              Event_manager.set_comp_values("cargando",f"preparandose para cargar los datos: {str(porcent)} %")
            else:
              Event_manager.set_comp_values("cargando","preparandose para cargar los datos: 0 %")
            tabla_name=content[1][i]
            filename=content[2]+tabla_name+".csv"
            conexion_bd.set_tabla(tabla_name)
            if(os.path.exists(filename)==False):
               cls.estatus=0
               cls.result=constantes.REQUEST_RESULT_ERROR_READ_CSV
               return constantes.REQUEST_RESULT_ERROR_READ_CSV
            if(conexion_bd.reset_table()==-1):
               cls.estatus=0
               cls.result=constantes.REQUEST_RESULT_ERROR_READ_CSV
               return constantes.REQUEST_RESULT_ERROR_READ_CSV
         conexion_bd.set_foregein_check(True)
         porcent=0 
         actual=0         
         for j in range(num_tablas-1,-1,-1):           
            if(actual>0):
              porcent=(int((float(actual/num_tablas))*100))
              Event_manager.set_comp_values("cargando",f"cargando los datos: {str(porcent)} %")
            else:
              Event_manager.set_comp_values("cargando","cargando los datos: 0 %")       
            tabla_name=content[1][j]
            filename=content[2]+tabla_name+".csv"
            conexion_bd.set_tabla(tabla_name)
            if(os.path.exists(filename)==False):        
               cls.estatus=0
               cls.result=constantes.REQUEST_RESULT_ERROR_READ_CSV
               return constantes.REQUEST_RESULT_ERROR_READ_CSV  
            file_tabla=open(filename, "r")
            data_linea=""
            count=0
            for linea in file_tabla:
              if(count==0):
                count+=1
              else:
                data_linea+=linea
                data_linea=data_linea.replace("\n","")
                data_temp=data_linea.split(';')
                data_linea=""
                if(data_temp!=[]):
                  conexion_bd.add_data(data_temp)
            file_tabla.close()
            actual=actual+1  
         cls.estatus=0
         cls.result=constantes.REQUEST_RESULT_READ_CSV_SUCCESS
         return constantes.REQUEST_RESULT_READ_CSV_SUCCESS    
            
